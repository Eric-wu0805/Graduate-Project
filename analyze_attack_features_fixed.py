import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def analyze_attack_features():
    """分析每個攻擊類型使用的特徵並輸出到Excel"""
    
    print("開始分析攻擊類型特徵...")
    
    # 創建Excel工作簿
    wb = Workbook()
    
    # 讀取KDD測試集和訓練集
    print("讀取KDD數據集...")
    df_test = pd.read_csv('kdd_test.csv')
    df_train = pd.read_csv('kdd_train.csv')
    
    # 合併數據集以獲得完整的攻擊類型列表
    df_combined = pd.concat([df_test, df_train], ignore_index=True)
    
    # 獲取所有實際的攻擊類型
    all_attack_types = sorted(df_combined['labels'].unique())
    print(f"發現的KDD攻擊類型: {all_attack_types}")
    print(f"總攻擊類型數量: {len(all_attack_types)}")
    
    # 定義特徵類別
    feature_categories = {
        '基本連接特徵': ['duration', 'protocol_type', 'service', 'flag', 'src_bytes', 'dst_bytes', 'land'],
        '內容特徵': ['wrong_fragment', 'urgent', 'hot', 'num_failed_logins', 'logged_in', 'num_compromised', 
                    'root_shell', 'su_attempted', 'num_root', 'num_file_creations', 'num_shells', 
                    'num_access_files', 'num_outbound_cmds', 'is_host_login', 'is_guest_login'],
        '流量統計特徵': ['count', 'srv_count', 'serror_rate', 'srv_serror_rate', 'rerror_rate', 
                       'srv_rerror_rate', 'same_srv_rate', 'diff_srv_rate', 'srv_diff_host_rate'],
        '目標主機統計特徵': ['dst_host_count', 'dst_host_srv_count', 'dst_host_same_srv_rate', 
                          'dst_host_diff_srv_rate', 'dst_host_same_src_port_rate', 
                          'dst_host_srv_diff_host_rate', 'dst_host_serror_rate', 
                          'dst_host_srv_serror_rate', 'dst_host_rerror_rate', 'dst_host_srv_rerror_rate']
    }
    
    # 創建KDD分析工作表
    ws_kdd = wb.active
    ws_kdd.title = "KDD攻擊特徵分析"
    
    # 設置標題
    ws_kdd['A1'] = "KDD Cup 1999 攻擊類型特徵分析"
    ws_kdd['A1'].font = Font(bold=True, size=16)
    ws_kdd.merge_cells('A1:I1')
    
    # 創建數據列表
    data_rows = []
    data_rows.append(['攻擊類型', '記錄數', '特徵類別', '特徵名稱', '使用率(%)', '最小值', '最大值', '均值', '說明'])
    
    # 分析每個攻擊類型
    for attack in all_attack_types:
        print(f"分析攻擊類型: {attack}")
        
        # 從測試集和訓練集中獲取該攻擊類型的數據
        attack_data_test = df_test[df_test['labels'] == attack]
        attack_data_train = df_train[df_train['labels'] == attack]
        
        # 合併測試集和訓練集的數據
        attack_data = pd.concat([attack_data_test, attack_data_train], ignore_index=True)
        record_count = len(attack_data)
        
        print(f"  - {attack}: 測試集 {len(attack_data_test)} 條, 訓練集 {len(attack_data_train)} 條, 總計 {record_count} 條")
        
        # 分析每個特徵類別
        for category, features in feature_categories.items():
            for feature in features:
                if feature in attack_data.columns:
                    if attack_data[feature].dtype in ['int64', 'float64']:
                        if feature in ['protocol_type', 'service', 'flag']:
                            # 分類特徵
                            unique_vals = attack_data[feature].unique()
                            unique_vals_str = str(list(unique_vals))[:50] + "..." if len(str(list(unique_vals))) > 50 else str(list(unique_vals))
                            data_rows.append([
                                attack, record_count, category, feature, 
                                '100%', 'N/A', 'N/A', 'N/A', 
                                f'分類特徵: {unique_vals_str}'
                            ])
                        else:
                            # 數值特徵
                            mean_val = attack_data[feature].mean()
                            min_val = attack_data[feature].min()
                            max_val = attack_data[feature].max()
                            non_zero_count = (attack_data[feature] != 0).sum()
                            usage_rate = non_zero_count / len(attack_data) * 100
                            
                            if usage_rate > 0:
                                data_rows.append([
                                    attack, record_count, category, feature,
                                    f'{usage_rate:.1f}%', f'{min_val:.2f}', f'{max_val:.2f}', f'{mean_val:.2f}',
                                    '已使用'
                                ])
                            else:
                                data_rows.append([
                                    attack, record_count, category, feature,
                                    '0%', '0', '0', '0',
                                    '未使用(全為0)'
                                ])
                    else:
                        # 分類特徵
                        unique_vals = attack_data[feature].unique()
                        unique_vals_str = str(list(unique_vals))[:50] + "..." if len(str(list(unique_vals))) > 50 else str(list(unique_vals))
                        data_rows.append([
                            attack, record_count, category, feature,
                            '100%', 'N/A', 'N/A', 'N/A',
                            f'分類特徵: {unique_vals_str}'
                        ])
    
    # 寫入KDD數據 - 分批寫入以避免內存問題
    print(f"寫入 {len(data_rows)} 行數據到Excel...")
    
    # 每1000行為一批
    batch_size = 1000
    for i in range(0, len(data_rows), batch_size):
        batch = data_rows[i:i+batch_size]
        for row in batch:
            ws_kdd.append(row)
        print(f"已寫入 {min(i+batch_size, len(data_rows))}/{len(data_rows)} 行")
    
    # 設置列寬
    ws_kdd.column_dimensions['A'].width = 20  # 攻擊類型
    ws_kdd.column_dimensions['B'].width = 12  # 記錄數
    ws_kdd.column_dimensions['C'].width = 25  # 特徵類別
    ws_kdd.column_dimensions['D'].width = 30  # 特徵名稱
    ws_kdd.column_dimensions['E'].width = 12  # 使用率
    ws_kdd.column_dimensions['F'].width = 12  # 最小值
    ws_kdd.column_dimensions['G'].width = 12  # 最大值
    ws_kdd.column_dimensions['H'].width = 12  # 均值
    ws_kdd.column_dimensions['I'].width = 40  # 說明
    
    # 設置標題行格式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_kdd[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # 創建UNSW-NB15分析工作表
    try:
        print("讀取UNSW-NB15數據集...")
        nb15_df = pd.read_csv('UNSW_NB15_testing-set.csv')
        
        ws_nb15 = wb.create_sheet("UNSW-NB15攻擊特徵分析")
        
        # 設置標題
        ws_nb15['A1'] = "UNSW-NB15 攻擊類型特徵分析"
        ws_nb15['A1'].font = Font(bold=True, size=16)
        ws_nb15.merge_cells('A1:I1')
        
        # 定義UNSW-NB15特徵類別
        nb15_feature_categories = {
            '基本連接特徵': ['dur', 'proto', 'service', 'state', 'srcip', 'sport', 'dstip', 'dsport'],
            '流量特徵': ['sbytes', 'dbytes', 'sttl', 'dttl', 'sloss', 'dloss', 'Sload', 'Dload', 'Spkts', 'Dpkts'],
            '時間特徵': ['Stime', 'Ltime', 'Sintpkt', 'Dintpkt', 'tcprtt', 'synack', 'ackdat'],
            '統計特徵': ['smeansz', 'dmeansz', 'trans_depth', 'res_bdy_len', 'Sjit', 'Djit'],
            '連接統計特徵': ['ct_srv_src', 'ct_state_ttl', 'ct_dst_ltm', 'ct_src_dport_ltm', 
                           'ct_dst_sport_ltm', 'ct_dst_src_ltm', 'ct_src_ltm', 'ct_srv_dst'],
            '協議特徵': ['is_ftp_login', 'ct_ftp_cmd', 'ct_flw_http_mthd', 'is_sm_ips_ports']
        }
        
        # 獲取所有攻擊類型
        nb15_attack_types = nb15_df['attack_cat'].unique()
        nb15_attack_types = sorted(nb15_attack_types)
        print(f"發現的UNSW-NB15攻擊類型: {nb15_attack_types}")
        
        # 創建數據列表
        nb15_data_rows = []
        nb15_data_rows.append(['攻擊類型', '記錄數', '特徵類別', '特徵名稱', '使用率(%)', '最小值', '最大值', '均值', '說明'])
        
        # 分析每個攻擊類型
        for attack in nb15_attack_types:
            print(f"分析UNSW-NB15攻擊類型: {attack}")
            attack_data = nb15_df[nb15_df['attack_cat'] == attack]
            record_count = len(attack_data)
            
            # 分析每個特徵類別
            for category, features in nb15_feature_categories.items():
                for feature in features:
                    if feature in attack_data.columns:
                        if attack_data[feature].dtype in ['int64', 'float64']:
                            if feature in ['proto', 'service', 'state', 'srcip', 'dstip']:
                                # 分類特徵
                                unique_count = attack_data[feature].nunique()
                                nb15_data_rows.append([
                                    attack, record_count, category, feature,
                                    '100%', 'N/A', 'N/A', 'N/A',
                                    f'分類特徵: {unique_count} 個唯一值'
                                ])
                            else:
                                # 數值特徵
                                mean_val = attack_data[feature].mean()
                                min_val = attack_data[feature].min()
                                max_val = attack_data[feature].max()
                                non_zero_count = (attack_data[feature] != 0).sum()
                                usage_rate = non_zero_count / len(attack_data) * 100
                                
                                if usage_rate > 0:
                                    nb15_data_rows.append([
                                        attack, record_count, category, feature,
                                        f'{usage_rate:.1f}%', f'{min_val:.2f}', f'{max_val:.2f}', f'{mean_val:.2f}',
                                        '已使用'
                                    ])
                                else:
                                    nb15_data_rows.append([
                                        attack, record_count, category, feature,
                                        '0%', '0', '0', '0',
                                        '未使用(全為0)'
                                    ])
                        else:
                            # 分類特徵
                            unique_count = attack_data[feature].nunique()
                            nb15_data_rows.append([
                                attack, record_count, category, feature,
                                '100%', 'N/A', 'N/A', 'N/A',
                                f'分類特徵: {unique_count} 個唯一值'
                            ])
        
        # 寫入UNSW-NB15數據 - 分批寫入
        print(f"寫入 {len(nb15_data_rows)} 行UNSW-NB15數據到Excel...")
        for i in range(0, len(nb15_data_rows), batch_size):
            batch = nb15_data_rows[i:i+batch_size]
            for row in batch:
                ws_nb15.append(row)
            print(f"已寫入 {min(i+batch_size, len(nb15_data_rows))}/{len(nb15_data_rows)} 行")
        
        # 設置列寬
        ws_nb15.column_dimensions['A'].width = 20
        ws_nb15.column_dimensions['B'].width = 12
        ws_nb15.column_dimensions['C'].width = 25
        ws_nb15.column_dimensions['D'].width = 30
        ws_nb15.column_dimensions['E'].width = 12
        ws_nb15.column_dimensions['F'].width = 12
        ws_nb15.column_dimensions['G'].width = 12
        ws_nb15.column_dimensions['H'].width = 12
        ws_nb15.column_dimensions['I'].width = 40
        
        # 設置標題行格式
        for cell in ws_nb15[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
    except FileNotFoundError:
        print("UNSW-NB15測試集文件未找到")
        nb15_attack_types = []
    
    # 創建統計摘要工作表
    ws_summary = wb.create_sheet("統計摘要")
    
    # 設置標題
    ws_summary['A1'] = "攻擊類型統計摘要"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:F1')
    
    # KDD統計摘要
    ws_summary['A3'] = "KDD Cup 1999 攻擊類型統計"
    ws_summary['A3'].font = Font(bold=True, size=14)
    
    kdd_summary_data = []
    kdd_summary_data.append(['攻擊類型', '測試集記錄數', '訓練集記錄數', '總記錄數', '百分比', '攻擊類別'])
    
    # 計算每個攻擊類型的統計
    for attack in all_attack_types:
        test_count = len(df_test[df_test['labels'] == attack])
        train_count = len(df_train[df_train['labels'] == attack])
        total_count = test_count + train_count
        percentage = total_count / len(df_combined) * 100
        
        # 判斷攻擊類別
        if attack == 'normal':
            attack_category = '正常流量'
        elif attack in ['neptune', 'smurf', 'back', 'teardrop', 'pod', 'land', 'apache2', 'processtable', 'mailbomb']:
            attack_category = 'DoS攻擊'
        elif attack in ['satan', 'ipsweep', 'portsweep', 'mscan', 'nmap', 'saint']:
            attack_category = '掃描攻擊'
        elif attack in ['guess_passwd', 'ftp_write', 'imap', 'phf', 'multihop', 'warezclient', 'warezmaster', 
                       'snmpguess', 'snmpgetattack', 'httptunnel', 'ps', 'sendmail', 'xlock', 'xterm', 
                       'named', 'perl', 'xsnoop']:
            attack_category = 'R2L攻擊'
        elif attack in ['buffer_overflow', 'loadmodule', 'rootkit']:
            attack_category = 'U2R攻擊'
        elif attack == 'spy':
            attack_category = '間諜攻擊'
        else:
            attack_category = '其他'
        
        kdd_summary_data.append([attack, test_count, train_count, total_count, f'{percentage:.2f}%', attack_category])
    
    # 寫入KDD摘要
    for row in kdd_summary_data:
        ws_summary.append(row)
    
    # UNSW-NB15統計摘要
    if nb15_attack_types:
        try:
            ws_summary['A' + str(len(kdd_summary_data) + 5)] = "UNSW-NB15 攻擊類型統計"
            ws_summary['A' + str(len(kdd_summary_data) + 5)].font = Font(bold=True, size=14)
            
            nb15_summary_data = []
            nb15_summary_data.append(['攻擊類型', '記錄數', '百分比', '攻擊類別'])
            
            for attack in nb15_attack_types:
                count = len(nb15_df[nb15_df['attack_cat'] == attack])
                percentage = count / len(nb15_df) * 100
                
                # 判斷攻擊類別
                if attack == 'Normal':
                    attack_category = '正常流量'
                elif attack == 'DoS':
                    attack_category = '拒絕服務攻擊'
                elif attack == 'Reconnaissance':
                    attack_category = '偵察攻擊'
                elif attack == 'Exploits':
                    attack_category = '漏洞利用攻擊'
                elif attack == 'Fuzzers':
                    attack_category = '模糊測試攻擊'
                elif attack == 'Analysis':
                    attack_category = '分析攻擊'
                elif attack == 'Backdoor':
                    attack_category = '後門攻擊'
                elif attack == 'Shellcode':
                    attack_category = 'Shellcode攻擊'
                elif attack == 'Worms':
                    attack_category = '蠕蟲攻擊'
                elif attack == 'Generic':
                    attack_category = '通用攻擊'
                else:
                    attack_category = '其他'
                
                nb15_summary_data.append([attack, count, f'{percentage:.2f}%', attack_category])
            
            # 寫入UNSW-NB15摘要
            start_row = len(kdd_summary_data) + 6
            for i, row in enumerate(nb15_summary_data):
                for j, value in enumerate(row):
                    ws_summary.cell(row=start_row + i, column=j + 1, value=value)
                    
        except:
            pass
    
    # 設置摘要工作表格式
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 20
    
    # 保存Excel文件
    print("保存Excel文件...")
    
    # 尝试不同的文件名，避免文件被占用
    base_filename = '攻擊類型特徵分析報告_完整版'
    counter = 1
    output_file = f'{base_filename}.xlsx'
    
    while os.path.exists(output_file) or os.path.exists(f'~${base_filename}.xlsx'):
        output_file = f'{base_filename}_{counter}.xlsx'
        counter += 1
        if counter > 100:  # 防止无限循环
            break
    
    try:
        wb.save(output_file)
        print(f"分析完成！結果已保存到 {output_file}")
    except PermissionError:
        print(f"無法保存到 {output_file}，文件可能被其他程序占用")
        print("請關閉Excel文件後重試，或檢查文件權限")
        return
    except Exception as e:
        print(f"保存文件時發生錯誤: {e}")
        return
    
    print(f"包含以下工作表:")
    print(f"1. KDD攻擊特徵分析 - KDD Cup 1999數據集的詳細特徵分析")
    if nb15_attack_types:
        print(f"2. UNSW-NB15攻擊特徵分析 - UNSW-NB15數據集的詳細特徵分析")
    print(f"3. 統計摘要 - 兩個數據集的攻擊類型統計摘要")
    print(f"\nKDD攻擊類型總數: {len(all_attack_types)}")
    if nb15_attack_types:
        print(f"UNSW-NB15攻擊類型總數: {len(nb15_attack_types)}")
    
    # 打印詳細的攻擊類型統計
    print(f"\n=== KDD攻擊類型詳細統計 ===")
    for attack in all_attack_types:
        test_count = len(df_test[df_test['labels'] == attack])
        train_count = len(df_train[df_train['labels'] == attack])
        total_count = test_count + train_count
        print(f"{attack}: 測試集 {test_count}, 訓練集 {train_count}, 總計 {total_count}")

if __name__ == "__main__":
    analyze_attack_features() 